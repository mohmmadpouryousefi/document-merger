"""
Excel merging functionality with formatting preservation.
"""

import logging
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook


class ExcelMerger:
    """Handles merging of Excel files while preserving formatting, formulas, and styles."""

    def __init__(self):
        """Initialize the Excel merger."""
        self.logger = logging.getLogger(__name__)

    def merge_excel_files(self, input_files: List[str], output_file: str) -> bool:
        """
        Merge multiple Excel files into a single Excel file with separate sheets.

        Args:
            input_files: List of input Excel file paths
            output_file: Output Excel file path

        Returns:
            True if successful, False otherwise
        """
        try:
            # Create a new workbook
            merged_workbook = Workbook()
            # Remove the default sheet
            merged_workbook.remove(merged_workbook.active)

            sheet_names = set()

            for i, excel_file in enumerate(input_files):
                try:
                    self.logger.info(f"Processing Excel file: {excel_file}")

                    # Load the source workbook
                    source_workbook = openpyxl.load_workbook(
                        excel_file, data_only=False
                    )

                    # Get a unique base name for sheets from this file
                    file_name = Path(excel_file).stem

                    # Copy each sheet from source to merged workbook
                    for sheet in source_workbook.worksheets:
                        # Create unique sheet name
                        original_name = sheet.title
                        new_name = f"{file_name}_{original_name}"

                        # Ensure sheet name is unique
                        counter = 1
                        while new_name in sheet_names:
                            new_name = f"{file_name}_{original_name}_{counter}"
                            counter += 1

                        sheet_names.add(new_name)

                        # Copy the sheet
                        new_sheet = merged_workbook.create_sheet(title=new_name)
                        self._copy_sheet(sheet, new_sheet)

                        self.logger.info(f"Copied sheet: {original_name} -> {new_name}")

                    source_workbook.close()

                except Exception as e:
                    self.logger.error(
                        f"Error processing Excel file {excel_file}: {str(e)}"
                    )
                    raise Exception(f"Failed to process {excel_file}: {str(e)}")

            # Save the merged workbook
            output_path = Path(output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            merged_workbook.save(output_file)
            merged_workbook.close()

            self.logger.info(
                f"Successfully merged {len(input_files)} Excel files into {output_file}"
            )
            return True

        except Exception as e:
            self.logger.error(f"Excel merge failed: {str(e)}")
            return False

    def _copy_sheet(self, source_sheet, target_sheet):
        """
        Copy all data, formatting, and styles from source sheet to target sheet.

        Args:
            source_sheet: Source worksheet
            target_sheet: Target worksheet
        """
        # Copy cell values and styles
        for row in source_sheet.iter_rows():
            for cell in row:
                if cell.value is not None or cell.has_style:
                    target_cell = target_sheet[cell.coordinate]

                    # Copy value
                    target_cell.value = cell.value

                    # Copy formatting
                    if cell.has_style:
                        target_cell.font = cell.font.copy()
                        target_cell.border = cell.border.copy()
                        target_cell.fill = cell.fill.copy()
                        target_cell.number_format = cell.number_format
                        target_cell.protection = cell.protection.copy()
                        target_cell.alignment = cell.alignment.copy()

        # Copy merged cells
        for merged_range in source_sheet.merged_cells.ranges:
            target_sheet.merge_cells(str(merged_range))

        # Copy column dimensions
        for col in source_sheet.column_dimensions:
            target_sheet.column_dimensions[col] = source_sheet.column_dimensions[col]

        # Copy row dimensions
        for row in source_sheet.row_dimensions:
            target_sheet.row_dimensions[row] = source_sheet.row_dimensions[row]

        # Copy sheet properties
        target_sheet.sheet_format = source_sheet.sheet_format
        target_sheet.sheet_properties = source_sheet.sheet_properties
        target_sheet.page_setup = source_sheet.page_setup
        target_sheet.print_options = source_sheet.print_options

    def validate_excel(self, file_path: str) -> bool:
        """
        Validate if an Excel file is readable and not corrupted.

        Args:
            file_path: Path to the Excel file

        Returns:
            True if valid, False otherwise
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)
            # Try to read the first sheet to check if Excel file is valid
            if workbook.worksheets:
                sheet = workbook.worksheets[0]
                # Try to access a cell to validate structure
                _ = sheet["A1"].value
            workbook.close()
            return True
        except Exception as e:
            self.logger.error(f"Excel validation failed for {file_path}: {str(e)}")
            return False

    def get_excel_info(self, file_path: str) -> dict:
        """
        Get information about an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            Dictionary with Excel file information
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            sheets_info = []
            for sheet in workbook.worksheets:
                sheet_info = {
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
                sheets_info.append(sheet_info)

            info = {
                "sheets": len(workbook.worksheets),
                "sheet_names": [sheet.title for sheet in workbook.worksheets],
                "sheets_info": sheets_info,
            }

            # Try to get workbook properties
            try:
                props = workbook.properties
                info.update(
                    {
                        "title": props.title,
                        "creator": props.creator,
                        "description": props.description,
                        "created": props.created.isoformat() if props.created else None,
                        "modified": (
                            props.modified.isoformat() if props.modified else None
                        ),
                    }
                )
            except:
                pass

            workbook.close()
            return info

        except Exception as e:
            self.logger.error(f"Failed to get Excel info for {file_path}: {str(e)}")
            return {"sheets": 0, "error": str(e)}

    def merge_specific_sheets(
        self, file_sheet_mapping: Dict[str, List[str]], output_file: str
    ) -> bool:
        """
        Merge specific sheets from multiple Excel files.

        Args:
            file_sheet_mapping: Dictionary mapping file paths to list of sheet names to include
            output_file: Output Excel file path

        Returns:
            True if successful, False otherwise
        """
        try:
            merged_workbook = Workbook()
            merged_workbook.remove(merged_workbook.active)

            sheet_names = set()

            for excel_file, sheet_names_to_copy in file_sheet_mapping.items():
                source_workbook = openpyxl.load_workbook(excel_file, data_only=False)
                file_name = Path(excel_file).stem

                for sheet_name in sheet_names_to_copy:
                    if sheet_name in [
                        sheet.title for sheet in source_workbook.worksheets
                    ]:
                        source_sheet = source_workbook[sheet_name]

                        # Create unique name
                        new_name = f"{file_name}_{sheet_name}"
                        counter = 1
                        while new_name in sheet_names:
                            new_name = f"{file_name}_{sheet_name}_{counter}"
                            counter += 1

                        sheet_names.add(new_name)

                        # Copy the sheet
                        new_sheet = merged_workbook.create_sheet(title=new_name)
                        self._copy_sheet(source_sheet, new_sheet)

                        self.logger.info(f"Copied sheet: {sheet_name} -> {new_name}")

                source_workbook.close()

            # Save the merged workbook
            output_path = Path(output_file)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            merged_workbook.save(output_file)
            merged_workbook.close()

            return True

        except Exception as e:
            self.logger.error(f"Specific sheet merge failed: {str(e)}")
            return False
