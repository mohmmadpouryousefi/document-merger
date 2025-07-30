"""
Integration tests for Document Merger application.
"""

import os
import shutil
import sys
import tempfile
import unittest
from pathlib import Path

# Add src to path for testing
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))


class TestDocumentMergerIntegration(unittest.TestCase):
    """Integration tests for the complete application."""

    def setUp(self):
        """Set up integration test fixtures."""
        self.test_dir = tempfile.mkdtemp()
        self.sample_dir = os.path.join(self.test_dir, "samples")
        os.makedirs(self.sample_dir, exist_ok=True)

    def tearDown(self):
        """Clean up integration test fixtures."""
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def create_sample_files(self):
        """Create sample files for testing."""
        try:
            # Try to create real PDF files for testing
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            # Create sample PDFs
            pdf1_path = os.path.join(self.sample_dir, "sample1.pdf")
            c1 = canvas.Canvas(pdf1_path, pagesize=letter)
            c1.setTitle("Sample Document 1")
            c1.drawString(100, 750, "This is Sample Document 1")
            c1.drawString(100, 730, "Page 1 content")
            c1.showPage()
            c1.drawString(100, 750, "This is page 2 of Sample Document 1")
            c1.showPage()
            c1.save()

            pdf2_path = os.path.join(self.sample_dir, "sample2.pdf")
            c2 = canvas.Canvas(pdf2_path, pagesize=letter)
            c2.setTitle("Sample Document 2")
            c2.drawString(100, 750, "This is Sample Document 2")
            c2.drawString(100, 730, "Different content here")
            c2.showPage()
            c2.save()

            return [pdf1_path, pdf2_path]

        except ImportError:
            # If reportlab not available, create minimal PDFs
            return self.create_minimal_pdfs()

    def create_minimal_pdfs(self):
        """Create minimal PDF files for testing when reportlab is not available."""
        # This is a very basic PDF structure
        minimal_pdf = b"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
>>
endobj

2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj

3 0 obj
<<
/Type /Page
/Parent 2 0 R
/MediaBox [0 0 612 792]
>>
endobj

xref
0 4
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
trailer
<<
/Size 4
/Root 1 0 R
>>
startxref
175
%%EOF"""

        pdf_files = []
        for i in range(1, 3):
            pdf_path = os.path.join(self.sample_dir, f"minimal{i}.pdf")
            with open(pdf_path, "wb") as f:
                f.write(minimal_pdf)
            pdf_files.append(pdf_path)

        return pdf_files

    def create_sample_excel_files(self):
        """Create sample Excel files for testing."""
        import openpyxl
        from openpyxl.workbook import Workbook

        # Create first Excel file
        excel1_path = os.path.join(self.sample_dir, "data1.xlsx")
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Sales_Q1"

        # Add headers
        headers1 = ["Product", "Units Sold", "Revenue"]
        for col, header in enumerate(headers1, 1):
            ws1.cell(row=1, column=col, value=header)

        # Add data
        data1 = [
            ["Product A", 100, 10000],
            ["Product B", 150, 22500],
            ["Product C", 75, 9375],
        ]
        for row, row_data in enumerate(data1, 2):
            for col, value in enumerate(row_data, 1):
                ws1.cell(row=row, column=col, value=value)

        wb1.save(excel1_path)

        # Create second Excel file
        excel2_path = os.path.join(self.sample_dir, "data2.xlsx")
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sales_Q2"

        # Add headers
        headers2 = ["Product", "Units Sold", "Revenue"]
        for col, header in enumerate(headers2, 1):
            ws2.cell(row=1, column=col, value=header)

        # Add data
        data2 = [
            ["Product A", 120, 12000],
            ["Product B", 180, 27000],
            ["Product C", 90, 11250],
            ["Product D", 60, 7200],
        ]
        for row, row_data in enumerate(data2, 2):
            for col, value in enumerate(row_data, 1):
                ws2.cell(row=row, column=col, value=value)

        wb2.save(excel2_path)

        return [excel1_path, excel2_path]

    def test_end_to_end_pdf_merge(self):
        """Test complete PDF merge workflow."""
        from core.file_merger import FileMerger

        # Create sample files
        pdf_files = self.create_sample_files()
        output_file = os.path.join(self.test_dir, "merged_output.pdf")

        # Initialize merger
        merger = FileMerger()

        # Validate files using preview_merge
        preview_result = merger.preview_merge(pdf_files)
        self.assertTrue(preview_result["valid"], "PDF validation should succeed")

        # Get file info for first file
        file_info = merger.get_file_info(pdf_files[0])
        self.assertGreater(len(file_info), 0, "Should get file information")

        # Merge files
        merge_result = merger.merge_files(pdf_files, output_file, add_bookmarks=True)
        self.assertTrue(
            merge_result["success"],
            f"PDF merge should succeed: {merge_result.get('message', '')}",
        )

        # Verify output file exists
        self.assertTrue(os.path.exists(output_file), "Output file should exist")

        # Verify output file is valid PDF
        try:
            import PyPDF2

            with open(output_file, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                self.assertGreater(len(reader.pages), 0, "Output should have pages")
        except ImportError:
            # If PyPDF2 not available, just check file size
            self.assertGreater(
                os.path.getsize(output_file), 0, "Output file should not be empty"
            )

    def test_end_to_end_excel_merge(self):
        """Test complete Excel merge workflow."""
        from core.file_merger import FileMerger

        # Create sample files
        excel_files = self.create_sample_excel_files()
        output_file = os.path.join(self.test_dir, "merged_data.xlsx")

        # Initialize merger
        merger = FileMerger()

        # Validate files using preview_merge
        preview_result = merger.preview_merge(excel_files)
        self.assertTrue(preview_result["valid"], "Excel validation should succeed")

        # Get file info for first file
        file_info = merger.get_file_info(excel_files[0])
        self.assertGreater(len(file_info), 0, "Should get file information")

        # Merge files
        merge_result = merger.merge_files(excel_files, output_file)
        self.assertTrue(
            merge_result["success"],
            f"Excel merge should succeed: {merge_result.get('message', '')}",
        )

        # Verify output file exists
        self.assertTrue(os.path.exists(output_file), "Output file should exist")

        # Verify output file is valid Excel
        import openpyxl

        merged_workbook = openpyxl.load_workbook(output_file)
        self.assertGreater(
            len(merged_workbook.sheetnames), 0, "Output should have sheets"
        )

    def test_mixed_file_types_error(self):
        """Test that mixing file types produces appropriate error."""
        from core.file_merger import FileMerger

        # Create different file types
        pdf_files = self.create_sample_files()
        excel_files = self.create_sample_excel_files()

        mixed_files = [pdf_files[0], excel_files[0]]
        output_file = os.path.join(self.test_dir, "mixed_output.pdf")

        merger = FileMerger()

        # This should fail
        result = merger.merge_files(mixed_files, output_file)
        self.assertFalse(result["success"], "Mixing file types should fail")
        self.assertIn(
            "must be of the same type",
            result["message"],
            "Should indicate mixed file types error",
        )

    def test_large_file_handling(self):
        """Test handling of larger files."""
        from core.file_merger import FileMerger

        # Create larger Excel files
        large_excel_files = []
        for i in range(2):
            excel_path = os.path.join(self.sample_dir, f"large_data_{i}.xlsx")

            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"LargeData_{i}"

            # Add headers
            headers = ["ID", "Name", "Value", "Category", "Date"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)

            # Add 1000 rows of data
            for row in range(2, 1002):
                ws.cell(row=row, column=1, value=row - 1)
                ws.cell(row=row, column=2, value=f"Item_{row-1}")
                ws.cell(row=row, column=3, value=(row - 1) * 10)
                ws.cell(row=row, column=4, value=f"Category_{(row-1) % 5}")
                ws.cell(row=row, column=5, value=f"2023-{(row-1) % 12 + 1:02d}-01")

            wb.save(excel_path)
            large_excel_files.append(excel_path)

        output_file = os.path.join(self.test_dir, "large_merged.xlsx")

        merger = FileMerger()
        result = merger.merge_files(large_excel_files, output_file)

        self.assertTrue(result, "Large file merge should succeed")
        self.assertTrue(os.path.exists(output_file), "Large output file should exist")

        # Verify merged content
        import openpyxl

        merged_wb = openpyxl.load_workbook(output_file)
        self.assertEqual(len(merged_wb.sheetnames), 2, "Should have 2 sheets")

    def test_error_recovery(self):
        """Test error recovery and partial success scenarios."""
        from core.file_merger import FileMerger

        # Create one valid and one invalid file
        valid_files = self.create_sample_files()

        # Create an invalid PDF file
        invalid_pdf = os.path.join(self.sample_dir, "invalid.pdf")
        with open(invalid_pdf, "w") as f:
            f.write("This is not a valid PDF file")

        mixed_valid_invalid = [valid_files[0], invalid_pdf]
        output_file = os.path.join(self.test_dir, "recovery_test.pdf")

        merger = FileMerger()

        # The merger should handle this gracefully - either succeed with warnings or fail appropriately
        merge_result = merger.merge_files(mixed_valid_invalid, output_file)

        # Either the merge succeeds with warnings about the corrupted file,
        # or it fails completely depending on the implementation strategy
        if merge_result["success"]:
            # If it succeeds, there should be warnings about the corrupted file
            self.assertGreater(
                len(merge_result.get("warnings", [])),
                0,
                "Should have warnings about corrupted files",
            )
            # And the output should contain only the valid file content
            self.assertTrue(
                os.path.exists(output_file), "Output file should be created"
            )
        else:
            # If it fails, there should be error messages explaining why
            self.assertGreater(
                len(merge_result.get("errors", [])),
                0,
                "Should have error messages explaining the failure",
            )

        # In either case, the invalid file should be detected
        log_messages = (
            merge_result.get("message", "")
            + " ".join(merge_result.get("warnings", []))
            + " ".join(merge_result.get("errors", []))
        )
        self.assertTrue(
            any(
                keyword in log_messages.lower()
                for keyword in ["corrupt", "invalid", "failed", "error", "warning"]
            ),
            f"Should indicate file problems in messages: {log_messages}",
        )

    def test_performance_benchmarking(self):
        """Test performance with timing measurements."""
        import time

        from core.file_merger import FileMerger

        # Create multiple files for performance testing
        pdf_files = []
        for i in range(5):  # Create 5 files
            try:
                from reportlab.pdfgen import canvas

                pdf_path = os.path.join(self.sample_dir, f"perf_test_{i}.pdf")
                c = canvas.Canvas(pdf_path)

                # Add multiple pages
                for page in range(3):
                    c.drawString(
                        100, 750, f"Performance test file {i}, page {page + 1}"
                    )
                    c.drawString(100, 730, f"Content for benchmarking")
                    c.showPage()

                c.save()
                pdf_files.append(pdf_path)

            except ImportError:
                # Use minimal PDFs if reportlab not available
                pdf_files = self.create_sample_files()
                break

        output_file = os.path.join(self.test_dir, "performance_test.pdf")

        merger = FileMerger()

        # Time the merge operation
        start_time = time.time()
        result = merger.merge_files(pdf_files, output_file, add_bookmarks=True)
        end_time = time.time()

        merge_duration = end_time - start_time

        self.assertTrue(result, "Performance test merge should succeed")
        self.assertTrue(
            os.path.exists(output_file), "Performance test output should exist"
        )

        # Log performance results (not an assertion, just informational)
        print(f"Merged {len(pdf_files)} files in {merge_duration:.2f} seconds")


if __name__ == "__main__":
    unittest.main()
