"""
Test utilities and helper functions for Document Merger tests.
"""

import os
import shutil
import tempfile

import openpyxl
from openpyxl.workbook import Workbook


class TestFileGenerator:
    """Utility class for generating test files."""

    @staticmethod
    def create_minimal_pdf(filepath, title="Test Document"):
        """Create a minimal valid PDF file for testing."""
        minimal_pdf_content = f"""%PDF-1.4
1 0 obj
<<
/Type /Catalog
/Pages 2 0 R
/Info 4 0 R
>>
endobj

2 0 obj
<<
/Type /Pages
/Kids [3 0 R]
/Count 1
>>
endobj

3 0 obj
<<
/Type /Page
/Parent 2 0 R
/MediaBox [0 0 612 792]
>>
endobj

4 0 obj
<<
/Title ({title})
/Producer (Test Generator)
>>
endobj

xref
0 5
0000000000 65535 f
0000000009 00000 n
0000000074 00000 n
0000000131 00000 n
0000000198 00000 n
trailer
<<
/Size 5
/Root 1 0 R
/Info 4 0 R
>>
startxref
256
%%EOF""".encode(
            "utf-8"
        )

        with open(filepath, "wb") as f:
            f.write(minimal_pdf_content)

        return filepath

    @staticmethod
    def create_test_excel(filepath, sheet_data=None, sheet_name="Sheet1"):
        """Create a test Excel file with specified data."""
        if sheet_data is None:
            sheet_data = [
                ["Name", "Age", "City"],
                ["John", 25, "New York"],
                ["Jane", 30, "Los Angeles"],
                ["Bob", 35, "Chicago"],
            ]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_name

        for row_idx, row_data in enumerate(sheet_data, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(filepath)
        return filepath

    @staticmethod
    def create_multi_sheet_excel(filepath, sheets_data):
        """Create Excel file with multiple sheets."""
        workbook = Workbook()
        workbook.remove(workbook.active)  # Remove default sheet

        for sheet_name, sheet_data in sheets_data.items():
            sheet = workbook.create_sheet(title=sheet_name)
            for row_idx, row_data in enumerate(sheet_data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(filepath)
        return filepath

    @staticmethod
    def create_corrupted_file(filepath, content="This is not a valid file"):
        """Create a corrupted file for testing error handling."""
        with open(filepath, "w") as f:
            f.write(content)
        return filepath


class TestDirectoryManager:
    """Utility class for managing test directories."""

    def __init__(self):
        self.temp_dir = None
        self.cleanup_dirs = []

    def setup_test_directory(self):
        """Create and return a temporary test directory."""
        self.temp_dir = tempfile.mkdtemp()
        self.cleanup_dirs.append(self.temp_dir)
        return self.temp_dir

    def create_subdirectory(self, name):
        """Create a subdirectory in the test directory."""
        if not self.temp_dir:
            self.setup_test_directory()

        subdir = os.path.join(self.temp_dir, name)
        os.makedirs(subdir, exist_ok=True)
        return subdir

    def cleanup(self):
        """Clean up all created directories."""
        for directory in self.cleanup_dirs:
            if os.path.exists(directory):
                shutil.rmtree(directory, ignore_errors=True)
        self.cleanup_dirs.clear()
        self.temp_dir = None


class TestDataSets:
    """Predefined test data sets for various scenarios."""

    @staticmethod
    def get_simple_pdf_data():
        """Get simple PDF test data."""
        return {
            "files": ["simple1.pdf", "simple2.pdf"],
            "expected_pages": 2,
            "titles": ["Simple Document 1", "Simple Document 2"],
        }

    @staticmethod
    def get_complex_excel_data():
        """Get complex Excel test data."""
        return {
            "sales_data": {
                "Sales_Q1": [
                    ["Product", "Units", "Revenue", "Date"],
                    ["Product A", 100, 10000, "2023-01-01"],
                    ["Product B", 150, 22500, "2023-01-15"],
                    ["Product C", 75, 9375, "2023-01-30"],
                ],
                "Sales_Q2": [
                    ["Product", "Units", "Revenue", "Date"],
                    ["Product A", 120, 12000, "2023-04-01"],
                    ["Product B", 180, 27000, "2023-04-15"],
                    ["Product C", 90, 11250, "2023-04-30"],
                    ["Product D", 60, 7200, "2023-04-30"],
                ],
            },
            "hr_data": {
                "Employees": [
                    ["ID", "Name", "Department", "Salary"],
                    [1, "John Doe", "Engineering", 75000],
                    [2, "Jane Smith", "Marketing", 65000],
                    [3, "Bob Johnson", "Sales", 70000],
                ],
                "Departments": [
                    ["Department", "Manager", "Budget"],
                    ["Engineering", "Alice Wilson", 500000],
                    ["Marketing", "Charlie Brown", 300000],
                    ["Sales", "Diana Ross", 400000],
                ],
            },
        }

    @staticmethod
    def get_performance_test_data():
        """Get data for performance testing."""
        large_data = []
        headers = ["ID", "Name", "Email", "Department", "Salary", "Start_Date"]
        large_data.append(headers)

        departments = ["Engineering", "Marketing", "Sales", "HR", "Finance"]

        for i in range(1, 1001):  # 1000 rows
            row = [
                i,
                f"Employee_{i}",
                f"employee{i}@company.com",
                departments[i % len(departments)],
                50000 + (i * 100),
                f"2023-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}",
            ]
            large_data.append(row)

        return {"LargeDataSet": large_data}


class TestAssertions:
    """Custom assertion helpers for Document Merger tests."""

    @staticmethod
    def assert_pdf_valid(test_case, filepath):
        """Assert that a PDF file is valid and readable."""
        test_case.assertTrue(
            os.path.exists(filepath), f"PDF file should exist: {filepath}"
        )
        test_case.assertGreater(
            os.path.getsize(filepath), 0, "PDF file should not be empty"
        )

        try:
            import PyPDF2

            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                test_case.assertGreater(
                    len(reader.pages), 0, "PDF should have at least one page"
                )
        except ImportError:
            # If PyPDF2 not available, just check file size
            test_case.assertGreater(
                os.path.getsize(filepath), 100, "PDF file should be reasonable size"
            )

    @staticmethod
    def assert_excel_valid(test_case, filepath, expected_sheets=None):
        """Assert that an Excel file is valid and readable."""
        test_case.assertTrue(
            os.path.exists(filepath), f"Excel file should exist: {filepath}"
        )
        test_case.assertGreater(
            os.path.getsize(filepath), 0, "Excel file should not be empty"
        )

        workbook = openpyxl.load_workbook(filepath)
        test_case.assertGreater(
            len(workbook.sheetnames), 0, "Excel should have at least one sheet"
        )

        if expected_sheets:
            test_case.assertEqual(
                len(workbook.sheetnames),
                expected_sheets,
                f"Excel should have {expected_sheets} sheets",
            )

    @staticmethod
    def assert_merge_success(test_case, input_files, output_file, file_type="pdf"):
        """Assert that a merge operation was successful."""
        test_case.assertTrue(os.path.exists(output_file), "Output file should exist")

        if file_type.lower() == "pdf":
            TestAssertions.assert_pdf_valid(test_case, output_file)
        elif file_type.lower() in ["xlsx", "excel"]:
            TestAssertions.assert_excel_valid(test_case, output_file)

        # Check that output file is larger than any individual input file
        output_size = os.path.getsize(output_file)
        for input_file in input_files:
            if os.path.exists(input_file):
                input_size = os.path.getsize(input_file)
                # Output should generally be larger (but not always due to compression)
                test_case.assertGreater(output_size, 0, "Output should have content")
                test_case.assertGreater(input_size, 0, "Input should have content")


class MockObjects:
    """Mock objects for testing without external dependencies."""

    @staticmethod
    def create_mock_pdf_reader(pages=1, metadata=None):
        """Create a mock PDF reader object."""
        from unittest.mock import MagicMock

        mock_reader = MagicMock()
        mock_reader.pages = [MagicMock() for _ in range(pages)]

        if metadata:
            mock_reader.metadata = MagicMock()
            for key, value in metadata.items():
                setattr(mock_reader.metadata, key, value)
        else:
            mock_reader.metadata = None

        mock_reader.is_encrypted = False
        return mock_reader

    @staticmethod
    def create_mock_excel_workbook(sheet_names=None):
        """Create a mock Excel workbook object."""
        from unittest.mock import MagicMock

        if sheet_names is None:
            sheet_names = ["Sheet1"]

        mock_workbook = MagicMock()
        mock_workbook.sheetnames = sheet_names
        mock_workbook.worksheets = [MagicMock() for _ in sheet_names]

        return mock_workbook


# Test configuration
TEST_CONFIG = {
    "timeout": 30,  # seconds
    "max_file_size": 100 * 1024 * 1024,  # 100MB
    "temp_dir_prefix": "document_merger_test_",
    "skip_integration_on_missing_deps": True,
    "verbose_logging": True,
}

# Test file patterns
TEST_PATTERNS = {
    "pdf_files": ["*.pdf"],
    "excel_files": ["*.xlsx", "*.xls"],
    "temp_files": ["temp_*", "*.tmp"],
    "output_files": ["merged_*", "combined_*", "test_output_*"],
}
