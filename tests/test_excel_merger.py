"""
Unit tests for Excel merger functionality.
"""

# Add src to path for testing
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))

import shutil
import tempfile
import unittest

import openpyxl
from openpyxl.workbook import Workbook

from core.excel_merger import ExcelMerger


class TestExcelMerger(unittest.TestCase):
    """Test cases for ExcelMerger class."""

    def setUp(self):
        """Set up test fixtures."""
        self.merger = ExcelMerger()
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up test fixtures."""
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def create_test_excel(self, filename, sheet_data=None):
        """Create a test Excel file with sample data."""
        if sheet_data is None:
            sheet_data = {"Sheet1": [["Name", "Age"], ["John", 25], ["Jane", 30]]}

        filepath = os.path.join(self.test_dir, filename)
        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        for sheet_name, data in sheet_data.items():
            sheet = workbook.create_sheet(title=sheet_name)
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(filepath)
        return filepath

    def test_merger_initialization(self):
        """Test ExcelMerger initialization."""
        merger = ExcelMerger()
        self.assertIsInstance(merger, ExcelMerger)
        self.assertIsNotNone(merger.logger)

    def test_validate_excel_valid_file(self):
        """Test Excel validation with valid file."""
        excel_file = self.create_test_excel("valid.xlsx")

        result = self.merger.validate_excel(excel_file)
        self.assertTrue(result)

    def test_validate_excel_invalid_file(self):
        """Test Excel validation with invalid file."""
        invalid_file = os.path.join(self.test_dir, "invalid.xlsx")
        with open(invalid_file, "w") as f:
            f.write("invalid content")

        result = self.merger.validate_excel(invalid_file)
        self.assertFalse(result)

    def test_validate_excel_nonexistent_file(self):
        """Test Excel validation with non-existent file."""
        nonexistent_file = os.path.join(self.test_dir, "nonexistent.xlsx")

        result = self.merger.validate_excel(nonexistent_file)
        self.assertFalse(result)

    def test_get_excel_info_success(self):
        """Test getting Excel info from valid file."""
        sheet_data = {
            "Sales": [["Product", "Revenue"], ["A", 100], ["B", 200]],
            "Marketing": [["Campaign", "Cost"], ["X", 50], ["Y", 75]],
        }
        excel_file = self.create_test_excel("info_test.xlsx", sheet_data)

        info = self.merger.get_excel_info(excel_file)

        self.assertEqual(info["sheets"], 2)
        self.assertIn("Sales", info["sheet_names"])
        self.assertIn("Marketing", info["sheet_names"])
        self.assertIn("sheets_info", info)
        self.assertIsNone(info.get("error"))

    def test_get_excel_info_error(self):
        """Test getting Excel info from corrupted file."""
        invalid_file = os.path.join(self.test_dir, "corrupted.xlsx")
        with open(invalid_file, "w") as f:
            f.write("corrupted data")

        info = self.merger.get_excel_info(invalid_file)

        self.assertEqual(info["sheets"], 0)
        self.assertIn("error", info)

    def test_merge_excel_files_success(self):
        """Test successful Excel file merging."""
        # Create test files
        excel1_data = {"Data1": [["A", "B"], [1, 2], [3, 4]]}
        excel2_data = {"Data2": [["X", "Y"], [10, 20], [30, 40]]}

        excel1 = self.create_test_excel("file1.xlsx", excel1_data)
        excel2 = self.create_test_excel("file2.xlsx", excel2_data)
        output_file = os.path.join(self.test_dir, "merged.xlsx")

        result = self.merger.merge_excel_files([excel1, excel2], output_file)

        self.assertTrue(result)
        self.assertTrue(os.path.exists(output_file))

        # Verify merged content
        merged_workbook = openpyxl.load_workbook(output_file)
        sheet_names = merged_workbook.sheetnames

        # Should have sheets from both files
        self.assertGreater(len(sheet_names), 0)

    def test_merge_excel_files_with_duplicate_sheet_names(self):
        """Test merging files with duplicate sheet names."""
        # Both files have 'Sheet1'
        excel1_data = {"Sheet1": [["A"], [1]]}
        excel2_data = {"Sheet1": [["B"], [2]]}

        excel1 = self.create_test_excel("dup1.xlsx", excel1_data)
        excel2 = self.create_test_excel("dup2.xlsx", excel2_data)
        output_file = os.path.join(self.test_dir, "merged_dup.xlsx")

        result = self.merger.merge_excel_files([excel1, excel2], output_file)

        self.assertTrue(result)

        # Verify both sheets exist with unique names
        merged_workbook = openpyxl.load_workbook(output_file)
        sheet_names = merged_workbook.sheetnames

        # Should have 2 sheets with different names
        self.assertEqual(len(sheet_names), 2)
        self.assertNotEqual(sheet_names[0], sheet_names[1])

    def test_merge_excel_files_file_not_found(self):
        """Test merging with non-existent file."""
        excel1 = self.create_test_excel("exists.xlsx")
        nonexistent = os.path.join(self.test_dir, "missing.xlsx")
        output_file = os.path.join(self.test_dir, "merged.xlsx")

        result = self.merger.merge_excel_files([excel1, nonexistent], output_file)

        self.assertFalse(result)

    def test_merge_excel_files_corrupted_file(self):
        """Test merging with corrupted Excel file."""
        excel1 = self.create_test_excel("good.xlsx")

        # Create corrupted file
        corrupted = os.path.join(self.test_dir, "corrupted.xlsx")
        with open(corrupted, "w") as f:
            f.write("not an excel file")

        output_file = os.path.join(self.test_dir, "merged.xlsx")

        result = self.merger.merge_excel_files([excel1, corrupted], output_file)

        self.assertFalse(result)

    def test_output_directory_creation(self):
        """Test that output directories are created automatically."""
        excel1 = self.create_test_excel("file1.xlsx")
        nested_output = os.path.join(self.test_dir, "subdir", "nested", "output.xlsx")

        result = self.merger.merge_excel_files([excel1], nested_output)

        self.assertTrue(result)
        self.assertTrue(os.path.exists(nested_output))
        self.assertTrue(os.path.exists(os.path.dirname(nested_output)))

    def test_merge_with_different_sheet_structures(self):
        """Test merging files with different sheet structures."""
        # File 1: Simple data
        excel1_data = {"Simple": [["Name"], ["Alice"]]}

        # File 2: Complex data with multiple columns
        excel2_data = {
            "Complex": [
                ["ID", "Name", "Age", "City"],
                [1, "Bob", 25, "NYC"],
                [2, "Carol", 30, "LA"],
            ]
        }

        excel1 = self.create_test_excel("simple.xlsx", excel1_data)
        excel2 = self.create_test_excel("complex.xlsx", excel2_data)
        output_file = os.path.join(self.test_dir, "mixed.xlsx")

        result = self.merger.merge_excel_files([excel1, excel2], output_file)

        self.assertTrue(result)

        # Verify both sheet structures are preserved
        merged_workbook = openpyxl.load_workbook(output_file)
        self.assertEqual(len(merged_workbook.sheetnames), 2)


class TestExcelMergerIntegration(unittest.TestCase):
    """Integration tests for ExcelMerger using real openpyxl operations."""

    def setUp(self):
        """Set up integration test fixtures."""
        self.merger = ExcelMerger()
        self.test_dir = tempfile.mkdtemp()

    def tearDown(self):
        """Clean up integration test fixtures."""
        shutil.rmtree(self.test_dir, ignore_errors=True)

    def create_test_excel(self, filename, sheet_data=None):
        """Create a test Excel file with sample data."""
        if sheet_data is None:
            sheet_data = {"Sheet1": [["Name", "Age"], ["John", 25], ["Jane", 30]]}

        filepath = os.path.join(self.test_dir, filename)
        workbook = Workbook()

        # Remove default sheet
        workbook.remove(workbook.active)

        for sheet_name, data in sheet_data.items():
            sheet = workbook.create_sheet(title=sheet_name)
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, cell_value in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        workbook.save(filepath)
        return filepath

    def create_realistic_excel(self, filename, department):
        """Create a realistic Excel file with formatted data."""
        filepath = os.path.join(self.test_dir, filename)
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = f"{department}_Data"

        # Headers
        headers = ["Employee ID", "Name", "Department", "Salary", "Start Date"]
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Sample data
        data = [
            [101, "John Doe", department, 75000, "2023-01-15"],
            [102, "Jane Smith", department, 82000, "2022-08-20"],
            [103, "Bob Johnson", department, 68000, "2023-03-10"],
        ]

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                sheet.cell(row=row_idx, column=col_idx, value=value)

        workbook.save(filepath)
        return filepath

    def test_realistic_department_merge(self):
        """Integration test with realistic department data."""
        # Create department files
        sales_file = self.create_realistic_excel("sales.xlsx", "Sales")
        marketing_file = self.create_realistic_excel("marketing.xlsx", "Marketing")
        hr_file = self.create_realistic_excel("hr.xlsx", "HR")

        output_file = os.path.join(self.test_dir, "company_data.xlsx")

        result = self.merger.merge_excel_files(
            [sales_file, marketing_file, hr_file], output_file
        )

        self.assertTrue(result)
        self.assertTrue(os.path.exists(output_file))

        # Verify merged content
        merged_workbook = openpyxl.load_workbook(output_file)

        # Should have 3 sheets
        self.assertEqual(len(merged_workbook.sheetnames), 3)

        # Verify data integrity
        for sheet_name in merged_workbook.sheetnames:
            sheet = merged_workbook[sheet_name]
            # Each sheet should have headers + 3 data rows
            self.assertGreaterEqual(sheet.max_row, 4)
            # Should have 5 columns
            self.assertEqual(sheet.max_column, 5)

    def test_large_dataset_merge(self):
        """Test merging with larger datasets."""
        # Create files with more data
        large_data1 = {"Large1": [["ID", "Value"]]}
        large_data2 = {"Large2": [["ID", "Score"]]}

        # Add 100 rows of data to each
        for i in range(1, 101):
            large_data1["Large1"].append([i, f"Value_{i}"])
            large_data2["Large2"].append([i, i * 10])

        excel1 = self.create_test_excel("large1.xlsx", large_data1)
        excel2 = self.create_test_excel("large2.xlsx", large_data2)
        output_file = os.path.join(self.test_dir, "large_merged.xlsx")

        result = self.merger.merge_excel_files([excel1, excel2], output_file)

        self.assertTrue(result)

        # Verify large data was merged correctly
        merged_workbook = openpyxl.load_workbook(output_file)

        for sheet_name in merged_workbook.sheetnames:
            sheet = merged_workbook[sheet_name]
            # Should have 101 rows (header + 100 data rows)
            self.assertEqual(sheet.max_row, 101)


if __name__ == "__main__":
    unittest.main()
